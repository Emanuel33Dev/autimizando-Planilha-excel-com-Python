import openpyxl
from openpyxl.styles import Font, Alignment
# nome da planilha
book = openpyxl.Workbook()
print(book.sheetnames)

# nome da página
book.create_sheet('Computadores')
pagina = book['Computadores']

# nome das colunas
pagina.append(['Eletrônica', 'Memória', 'Preço'])

# dados
pagina.append(['Computadores 1', '8gb Ram', 'R$2500'])
pagina.append(['Computadores 2', '16gb Ram', 'R$5500'])
pagina.append(['Computadores 3', '32gb Ram', 'R$8500'])
# Estilo e Tamanho da Fonte
for cell in pagina[1]:
    cell.font = Font(name='Tahoma', size=18, bold=True)

# Alinhando o texto das células ao centro
for row in pagina.iter_rows(min_row=2, max_row=pagina.max_row, min_col=1, max_col=len(pagina[1])):
    for cell in row:
        cell.alignment = Alignment(horizontal='center')


# Ajustar a largurar das Colunas
for column_cells in pagina.columns:
    lengsth = max(len(str(cell.value)) for cell in column_cells)
    pagina.column_dimensions[column_cells[0].column_letter].width = lengsth + 10

# Salvando a Planilha
book.save('Meus Computadores.xlsx')
